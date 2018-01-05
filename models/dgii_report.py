# -*- coding: utf-8 -*-
###############################################################################
#  Copyright (c) 2015 - Marcos Organizador de Negocios SRL.
#  (<https://marcos.do/>)
#  Write by Eneldo Serrata (eneldo@marcos.do)
#  See LICENSE file for full copyright and licensing details.
#
# Odoo Proprietary License v1.0
#
# This software and associated files (the "Software") may only be used
# (nobody can redistribute (or sell) your module once they have bought it,
# unless you gave them your consent)
# if you have purchased a valid license
# from the authors, typically via Odoo Apps, or if you have received a written
# agreement from the authors of the Software (see the COPYRIGHT file).
#
# You may develop Odoo modules that use the Software as a library (typically
# by depending on it, importing it and using its resources), but without
# copying any source code or material from the Software. You may distribute
# those modules under the license of your choice, provided that this license is
# compatible with the terms of the Odoo Proprietary License (For example:
# LGPL, MIT, or proprietary licenses similar to this one).
#
# It is forbidden to publish, distribute, sublicense, or sell copies of the
# Softwar or modified copies of the Software.
#
# The above copyright notice and this permission notice must be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
# DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR
# OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE
# USE OR OTHER DEALINGS IN THE SOFTWARE.
###############################################################################

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError

from openpyxl import load_workbook
import base64
import os

import re
import calendar
import io

import logging

_logger = logging.getLogger(__name__)


class DgiiReport(models.Model):
    _name = "dgii.report"
    _inherit = ['mail.thread']

    @api.multi
    @api.depends("purchase_report")
    def _purchase_report_totals(self):
        for rec in self:
            rec.ITBIS_TOTAL = 0
            rec.ITBIS_TOTAL_NC = 0
            rec.ITBIS_TOTAL_PAYMENT = 0

            rec.TOTAL_MONTO_FACTURADO = 0
            rec.TOTAL_MONTO_NC = 0
            rec.TOTAL_MONTO_PAYMENT = 0

            rec.ITBIS_RETENIDO = 0
            rec.RETENCION_RENTA = 0

            for purchase in rec.purchase_report:
                if purchase.NUMERO_COMPROBANTE_MODIFICADO is not False:
                    rec.ITBIS_TOTAL_NC += purchase.ITBIS_FACTURADO
                    rec.TOTAL_MONTO_NC += purchase.MONTO_FACTURADO
                    rec.RETENCION_RENTA -= purchase.RETENCION_RENTA
                    rec.ITBIS_RETENIDO -= purchase.ITBIS_RETENIDO
                else:
                    rec.ITBIS_TOTAL += purchase.ITBIS_FACTURADO
                    rec.TOTAL_MONTO_FACTURADO += purchase.MONTO_FACTURADO
                    rec.RETENCION_RENTA += purchase.RETENCION_RENTA
                    rec.ITBIS_RETENIDO += purchase.ITBIS_RETENIDO

            rec.ITBIS_TOTAL_PAYMENT = rec.ITBIS_TOTAL - rec.ITBIS_TOTAL_NC
            rec.TOTAL_MONTO_PAYMENT = rec.TOTAL_MONTO_FACTURADO - rec.TOTAL_MONTO_NC

    @api.multi
    @api.depends("sale_report")
    def _sale_report_totals(self):
        for rec in self:
            rec.SALE_ITBIS_TOTAL = 0
            rec.SALE_ITBIS_NC = 0
            rec.SALE_ITBIS_CHARGED = 0
            rec.SALE_TOTAL_MONTO_FACTURADO = 0
            rec.SALE_TOTAL_MONTO_NC = 0
            rec.SALE_TOTAL_MONTO_CHARGED = 0

            for sale in rec.sale_report:
                if sale.NUMERO_COMPROBANTE_MODIFICADO:
                    rec.SALE_ITBIS_NC += abs(sale.ITBIS_FACTURADO)
                    rec.SALE_TOTAL_MONTO_NC += abs(sale.MONTO_FACTURADO)
                else:
                    rec.SALE_ITBIS_TOTAL += sale.ITBIS_FACTURADO
                    rec.SALE_TOTAL_MONTO_FACTURADO += sale.MONTO_FACTURADO

            rec.SALE_ITBIS_CHARGED = rec.SALE_ITBIS_TOTAL - abs(rec.SALE_ITBIS_NC)
            rec.SALE_TOTAL_MONTO_CHARGED = rec.SALE_TOTAL_MONTO_FACTURADO - abs(rec.SALE_TOTAL_MONTO_NC)

    @api.multi
    @api.depends("purchase_report", "sale_report")
    def _count_records(self):
        for rec in self:
            rec.COMPRAS_CANTIDAD_REGISTRO = rec.purchase_report and len(
                rec.purchase_report)
            rec.VENTAS_CANTIDAD_REGISTRO = rec.sale_report and len(
                rec.sale_report)
            rec.CANCEL_CANTIDAD_REGISTRO = rec.cancel_report and len(
                rec.cancel_report)
            rec.EXTERIOR_CANTIDAD_REGISTRO = rec.exterior_filename and len(
                rec.exterior_report)

    company_id = fields.Many2one('res.company', 'Company', required=False,
                                 default=lambda self: self.env.user.company_id)
    name = fields.Char(string="Periodo MM/YYYY", required=True, size=7)
    positive_balance = fields.Float("SALDO A FAVOR ANTERIOR", required=True)

    it_filename = fields.Char()
    it_binary = fields.Binary(string="Archivo excel IT-1")

    ir17_filename = fields.Char()
    ir17_binary = fields.Binary(string="Archivo excel IR-17")

    # 606
    COMPRAS_CANTIDAD_REGISTRO = fields.Integer("Cantidad de registros",
                                               compute=_count_records)

    TOTAL_MONTO_FACTURADO = fields.Float("Monto Compras",
                                         compute=_purchase_report_totals)
    ITBIS_TOTAL = fields.Float("ITBIS Compras",
                               compute=_purchase_report_totals)

    TOTAL_MONTO_NC = fields.Float(u"Notas de Crédito",
                                  compute=_purchase_report_totals)
    ITBIS_TOTAL_NC = fields.Float(u"ITBIS Notas de Crédito",
                                  compute=_purchase_report_totals)

    TOTAL_MONTO_PAYMENT = fields.Float("Total Monto Facturado",
                                       compute=_purchase_report_totals)
    ITBIS_TOTAL_PAYMENT = fields.Float("ITBIS Facturado",
                                       compute=_purchase_report_totals)

    ITBIS_RETENIDO = fields.Float("ITBIS Retenido",
                                  compute=_purchase_report_totals)
    RETENCION_RENTA = fields.Float(u"Retención Renta",
                                   compute=_purchase_report_totals)

    purchase_report = fields.One2many("dgii.report.purchase.line",
                                      "dgii_report_id")
    purchase_filename = fields.Char()
    purchase_binary = fields.Binary(string="Archivo 606 TXT")

    # 607
    VENTAS_CANTIDAD_REGISTRO = fields.Integer("Cantidad de registros",
                                              compute=_count_records)

    SALE_TOTAL_MONTO_FACTURADO = fields.Float("Monto Ventas",
                                              compute=_sale_report_totals)
    SALE_ITBIS_TOTAL = fields.Float("ITBIS Ventas",
                                    compute=_sale_report_totals)

    SALE_TOTAL_MONTO_NC = fields.Float(u"Notas de Crédito",
                                       compute=_sale_report_totals)
    SALE_ITBIS_NC = fields.Float(u"ITBIS Notas de Crédito",
                                 compute=_sale_report_totals)

    SALE_TOTAL_MONTO_CHARGED = fields.Float("Total Monto Facturado",
                                            compute=_sale_report_totals)
    SALE_ITBIS_CHARGED = fields.Float("ITBIS Facturado",
                                      compute=_sale_report_totals)

    sale_report = fields.One2many("dgii.report.sale.line", "dgii_report_id")
    sale_filename = fields.Char()
    sale_binary = fields.Binary(string="Archivo 607 TXT")

    # 608
    CANCEL_CANTIDAD_REGISTRO = fields.Integer("Cantidad de registros",
                                              compute=_count_records)
    cancel_report = fields.One2many("dgii.cancel.report.line",
                                    "dgii_report_id")
    cancel_filename = fields.Char()
    cancel_binary = fields.Binary(string="Archivo 608 TXT")

    # 609
    EXTERIOR_CANTIDAD_REGISTRO = fields.Integer("Cantidad de registros",
                                                compute=_count_records)
    EXTERIOR_TOTAL_MONTO_FACTURADO = fields.Float()
    exterior_report = fields.One2many("dgii.exterior.report.line",
                                      "dgii_report_id")
    exterior_filename = fields.Char("Total Monto Facturado")
    exterior_binary = fields.Binary(string="Archivo 609 TXT")

    state = fields.Selection(
        [('draft', 'Nuevo'),
         ('error', 'Con errores'),
         ('done', 'Validado')],
        default="draft")

    @api.multi
    def generate_report(self):
        api_marcos = self.env["marcos.api.tools"]
        try:
            month, year = self.name.split("/")
            last_day = calendar.monthrange(int(year), int(month))[1]
            start_date = "{}-{}-01".format(year, month)
            end_date = "{}-{}-{}".format(year, month, last_day)
        except:
            raise ValidationError(_(u"Período inválido"))

        error_list = {}
        xls_dict = {"it1": {}, "ir17": {}}
        purchase_report = []
        sale_report = []
        cancel_report = []
        ext_report = []
        sale_line = 1
        purchase_line = 1
        cancel_line = 1
        ext_line = 1

        invoice_ids = self.env["account.invoice"].search([
            ('date_invoice', '>=', start_date),
            ('date_invoice', '<=', end_date)])

        fiscal_invoices = invoice_ids.filtered(
            lambda inv: inv.state in ('open', 'paid', 'cancel') and
            (inv.journal_id.purchase_type != "others" or
             inv.journal_id.ncf_control is not False)
            )

        paid_invoices = self.env["account.payment"].search(
            [('payment_date', '>=', start_date),
             ('payment_date', '<=', end_date),
             ('invoice_ids', '!=', False)])

        for paid_line in paid_invoices:
            fiscal_invoices |= paid_line.invoice_ids.filtered(
                lambda r: r.state == 'paid' and r.tax_line_ids.filtered(
                    lambda t: t.tax_id.purchase_tax_type in ('ritbis', 'isr')))

        draft_invoice_ids_set = invoice_ids.filtered(lambda x:
                                                      x.state == "draft")

        for draft_line in draft_invoice_ids_set:
            if not error_list.get(draft_line.id, False):
                error_list.update(
                    {draft_line.id: [(draft_line.type,
                                      draft_line.number,
                                      "Factura sin validar")]})
            else:
                error_list[draft_line.id].append(
                    (draft_line.type, draft_line.number,
                     "Factura sin validar"))

        for invoice in fiscal_invoices:
            if invoice.type in ("in_invoice", "in_refund") and invoice.journal_id.purchase_type in ("import", "others"):
                continue
            if not invoice.move_name:
                error_list[invoice.id].append(
                    (invoice.type, invoice.number,
                     "Factura sin NCF asignado."))
                continue

            TIPO_IDENTIFICACION = "3"
            RNC_CEDULA = invoice.partner_id.vat and re.sub(
                "[^0-9]", "", invoice.partner_id.vat.strip()) or False
            if RNC_CEDULA:
                if len(RNC_CEDULA) == 9:
                    TIPO_IDENTIFICACION = "1"
                elif len(RNC_CEDULA) == 11:
                    TIPO_IDENTIFICACION = "2"
                else:
                    RNC_CEDULA = ""

            NUMERO_COMPROBANTE_FISCAL = invoice.number
            NUMERO_COMPROBANTE_MODIFICADO, FACTURA_AFECTADA = False, False
            FECHA_PAGO = False
            FECHA_COMPROBANTE = invoice.date_invoice

            # ITBIS_RETENIDO, RETENCION_RENTA = 0, 0
            # if invoice.state == "paid":
            #     move_id = self.env["account.move.line"].search(
            #         [("move_id", "=", invoice.move_id.id),
            #          ('full_reconcile_id', '!=', False)])
            #     if invoice.journal_id.purchase_type:
            #         if move_id:
            #             retentions = self.env["account.move.line"].search(
            #                 [('invoice', '=', invoice.id),
            #                  ('payment_id', '!=', False),
            #                  ('tax_line_id', '!=', False)])
            #             if retentions:
            #                 for retention in retentions:
            #                     if retention.tax_line_id.purchase_tax_type == "ritbis":
            #                         ITBIS_RETENIDO += retention.credit
            #                     elif retention.tax_line_id.purchase_tax_type == "isr":
            #                         RETENCION_RENTA += retention.credit
            #                 FECHA_PAGO = retentions[0].date if move_id and move_id[0].date else False

            if (invoice.state != "cancel"
                and (invoice.journal_id.ncf_remote_validation
                     or invoice.journal_id.ncf_control)):
                if invoice.type in ("out_invoice", "out_refund",
                                       "in_invoice", "in_refund"):
                    if not api_marcos.is_identification(
                        invoice.partner_id.vat) and invoice.partner_id.sale_fiscal_type in ("fiscal", "gov", "special"):
                        error_msg = u"RNC/Cédula no es válida"
                        if not error_list.get(invoice.id, False):
                            error_list.update(
                                {invoice.id: [(invoice.type, invoice.number,
                                               error_msg)]})
                        else:
                            error_list[invoice.id].append(
                                (invoice.type, invoice.number,
                                 error_msg))
                        if not api_marcos.is_ncf(invoice.number, invoice.type):
                            error_msg = u"NCF no es válido"
                            if not error_list.get(invoice.id, False):
                                error_list.update({invoice.id: [
                                    (invoice.type, invoice.number,
                                     error_msg)]})
                            else:
                                error_list[invoice.id].append(
                                    (invoice.type, invoice.number,
                                     error_msg))
                        continue

                    if invoice.type in ("out_refund", "in_refund"):
                        NUMERO_COMPROBANTE_MODIFICADO_ID = invoice.search(
                            [('number', '=', invoice.origin)], limit=1
                            )

                        if not NUMERO_COMPROBANTE_MODIFICADO_ID:
                            error_msg = "Falta el Comprobante que Afecta"
                            if not error_list.get(invoice.id, False):
                                error_list.update({invoice.id: [
                                    (invoice.type, invoice.number,
                                     error_msg)]})
                            else:
                                error_list[invoice.id].append(
                                    (invoice.type, invoice.number, error_msg))
                        else:
                            NUMERO_COMPROBANTE_MODIFICADO = NUMERO_COMPROBANTE_MODIFICADO_ID.number
                            FACTURA_AFECTADA = NUMERO_COMPROBANTE_MODIFICADO_ID.id

                    if not invoice.number:
                        error_msg = "Factura validada con error"
                        if not error_list.get(invoice.id, False):
                            error_list.update({invoice.id: [
                                (invoice.type, invoice.number, error_msg)]})
                        else:
                            error_list[invoice.id].append(
                                (invoice.type, invoice.number, error_msg))

                if not invoice.number:
                    error_msg = "Factura validada con error"
                    if not error_list.get(invoice.id, False):
                        error_list.update({invoice.id: [
                            (invoice.type, invoice.number, error_msg)]})
                    else:
                        error_list[invoice.id].append(
                            (invoice.type, invoice.number, error_msg))

            commun_data = {
                "RNC_CEDULA": RNC_CEDULA,
                "TIPO_IDENTIFICACION": TIPO_IDENTIFICACION,
                "NUMERO_COMPROBANTE_FISCAL": NUMERO_COMPROBANTE_FISCAL,
                "NUMERO_COMPROBANTE_MODIFICADO": NUMERO_COMPROBANTE_MODIFICADO,
                "FECHA_COMPROBANTE": FECHA_COMPROBANTE,
                "FECHA_PAGO": FECHA_PAGO,
                "invoice": invoice.id,
                "inv_partner": invoice.partner_id.id,
                "MONTO_FACTURADO": False,
                "affected_nvoice_id": FACTURA_AFECTADA,
                "nc": True if FACTURA_AFECTADA else False,
                "ITBIS_RETENIDO": 0,
                "RETENCION_RENTA": 0
            }

            commun_data["MONTO_FACTURADO"] += abs(invoice.amount_untaxed_signed)

            for line in invoice.invoice_line_ids:
                taxes = line.invoice_line_tax_ids
                # if not taxes:
                #     if invoice.type in ("out_invoice", "out_refund"):
                #         line.write({"invoice_line_tax_ids": [
                #             (4, self.env.ref("l10n_do.{}_tax_0_sale".format(
                #                 self.company_id.id)).id, False)]})
                #     else:
                #         line.write({"invoice_line_tax_ids": [
                #             (4, self.env.ref("l10n_do.{}_tax_0_purch".format(
                #                 self.company_id.id)).id, False)]})
                #     taxes = line.invoice_line_tax_ids

                move_line_ids = self.env["account.move.line"].search(
                    [('move_id', '=', invoice.move_id.id),
                     ('name', '=', line.name)])

                for tax in taxes:
                    if tax.type_tax_use in ("purchase", "sale"): # and tax.tax_group_id.name == 'ITBIS':
                        for base_line in move_line_ids:
                            base_amount = abs(base_line.debit -
                                              base_line.credit)
                            if tax.base_it1_cels:
                                xls_cels = tax.base_it1_cels.split(",")
                                for xls_cel in xls_cels:
                                    if not xls_dict["it1"].get(xls_cel, False):
                                        xls_dict["it1"].update(
                                            {xls_cel: base_amount})
                                    else:
                                        xls_dict["it1"][xls_cel] += base_amount
                            if tax.base_ir17_cels:
                                xls_cels = tax.base_ir17_cels.split(",")
                                for xls_cel in xls_cels:
                                    xls_cel = xls_cel.split("%")
                                    if len(xls_cel) == 1:
                                        if not xls_dict["ir17"].get(xls_cel[0], False):
                                            xls_dict["ir17"].update(
                                                {xls_cel[0]: base_amount})
                                        else:
                                            xls_dict["ir17"][xls_cel[0]] += base_amount
                                    elif len(xls_cel) == 2:
                                        if not xls_dict["ir17"].get(xls_cel[0], False):
                                            xls_dict["ir17"].update(
                                                {xls_cel[0]: round(base_amount * (float(xls_cel[1]) / 100), 2)})
                                        else:
                                            xls_dict["ir17"][xls_cel[0]] += round(
                                                base_amount * (float(xls_cel[1]) / 100), 2)

            if invoice.move_id:
                taxes = []
                for tax_line in invoice.tax_line_ids:
                    tax = self.env["account.tax"].search(
                        [('name', '=', tax_line.name),
                         ('account_id', '=', tax_line.account_id.id)])
                    if not tax:
                        tax = self.env["account.tax"].search(
                            [('name', '=', tax_line.name)])
                    if tax:
                        taxes.append(tax)

            ITBIS_FACTURADO = 0.0
            ITBIS_RETENIDO, RETENCION_RENTA = 0, 0
            for tax in invoice.tax_line_ids:
                if invoice.currency_id != invoice.company_id.currency_id and tax.tax_id.tax_group_id.name == 'ITBIS':
                    currency_id = invoice.currency_id.with_context(
                        date=invoice.date_invoice)
                    ITBIS_FACTURADO += currency_id.compute(
                        abs(tax.amount), invoice.company_id.currency_id)
                elif tax.tax_id.tax_group_id.name == 'ITBIS':
                    ITBIS_FACTURADO += abs(tax.amount)
                if invoice.state == "paid":
                    if tax.tax_id.purchase_tax_type == "ritbis":
                        if invoice.currency_id == invoice.company_id.currency_id:
                            ITBIS_RETENIDO += abs(tax.amount)
                        else:
                            currency_id = invoice.currency_id.with_context(
                                date=invoice.date_invoice)
                            ITBIS_RETENIDO += currency_id.compute(
                                abs(tax.amount), invoice.company_id.currency_id)

                    elif tax.tax_id.purchase_tax_type == "isr":
                        if invoice.currency_id == invoice.company_id.currency_id:
                            RETENCION_RENTA += abs(tax.amount)
                        else:
                            currency_id = invoice.currency_id.with_context(
                                date=invoice.date_invoice)
                            RETENCION_RENTA += currency_id.compute(
                                abs(tax.amount), invoice.company_id.currency_id)
                    if RETENCION_RENTA > 0 or ITBIS_RETENIDO > 0:
                        FECHA_PAGO = invoice.date_invoice
                        for move_payment in invoice.payment_move_line_ids:
                            if FECHA_PAGO < move_payment.date:
                                FECHA_PAGO = move_payment.date
                        commun_data.update({"FECHA_PAGO": FECHA_PAGO})

                # TODO refactoring IR17 and IT1 DGII Report
                # if tax.tax_it1_cels:
                #     xls_cels = tax.tax_it1_cels.split(",")
                #     for xls_cel in xls_cels:
                #         if not xls_dict["it1"].get(xls_cel, False):
                #             xls_dict["it1"].update({xls_cel: amount})
                #         else:
                #             xls_dict["it1"][xls_cel] += amount
                #
                # if tax.tax_ir17_cels:
                #     xls_cels = tax.tax_ir17_cels.split(",")
                #     for xls_cel in xls_cels:
                #         if not xls_dict["ir17"].get(xls_cel, False):
                #             xls_dict["ir17"].update({xls_cel: amount})
                #         else:
                #             xls_dict["ir17"][xls_cel] += amount

            commun_data.update({"ITBIS_FACTURADO": ITBIS_FACTURADO})
            commun_data.update({"ITBIS_RETENIDO": ITBIS_RETENIDO})
            commun_data.update({"RETENCION_RENTA": RETENCION_RENTA})

            if invoice.type in ("out_invoice", "out_refund") and invoice.state != "cancel":
                commun_data.update({"LINE": sale_line})
                sale_report.append(commun_data)
                sale_line += 1
            elif invoice.type in ("out_invoice", "out_refund") and invoice.state == "cancel":
                commun_data.update({"LINE": cancel_line,
                                    "TIPO_ANULACION": invoice.anulation_type,
                                    "NUMERO_COMPROBANTE_FISCAL": invoice.move_name})
                cancel_report.append(commun_data)
                cancel_line += 1
            elif invoice.type in ("in_invoice", "in_refund") and invoice.state != "cancel" and invoice.journal_id.purchase_type == "exterior":
                commun_data.update({"LINE": ext_line})
                ext_report.append(commun_data)
                ext_line += 1
            elif invoice.type in ("in_invoice", "in_refund") and invoice.state != "cancel":
                commun_data.update(
                    {"LINE": purchase_line,
                     "TIPO_BIENES_SERVICIOS_COMPRADOS": invoice.expense_type if
                     invoice.expense_type else '09'}
                     )
                purchase_report.append(commun_data)
                purchase_line += 1
            else:
                continue
        error_ids = error_list.keys()

        # 607
        self.sale_report.unlink()
        new_sale_lines = []
        for sale in sale_report:
            if not sale["invoice"] in error_ids:
                new_sale_lines.append((0, 0, sale))
        self.write({"sale_report": new_sale_lines})

        # 608
        self.cancel_report.unlink()
        new_cancel_report = []
        for cancel in cancel_report:
            if not cancel["invoice"] in error_ids:
                new_cancel_report.append((0, 0, cancel))
        self.write({"cancel_report": new_cancel_report})

        # 609
        new_ext_report = []
        self.exterior_report.unlink()
        for ext in ext_report:
            if not ext["invoice"] in error_ids:
                new_ext_report.append((0, 0, ext))
        self.write({"exterior_report": new_ext_report})

        # 606
        new_purchase_lines = []
        self.purchase_report.unlink()
        for purchase in purchase_report:
            if not purchase["invoice"] in error_ids:
                new_purchase_lines.append((0, 0, purchase))
        self.write({"purchase_report": new_purchase_lines})

        self.generate_txt()

        # fill IT-1 excel file
        cwf = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "IT-1-2017.xlsx")
        wb = load_workbook(cwf)
        ws1 = wb.get_sheet_by_name("IT-1")  # Get sheet 1 in writeable copy
        xls_dict["it1"].update({"S43": self.positive_balance})
        for k, v in xls_dict["it1"].iteritems():
            ws1[k] = v

        period = self.name.split("/")
        FILENAME = "IT-1-{}-{}.xlsx".format(period[0], period[1])
        wb.save("/tmp/{}".format(FILENAME))
        with open("/tmp/{}".format(FILENAME), "rb") as xls_file:
            self.write({
                'it_filename': FILENAME,
                'it_binary': base64.b64encode(xls_file.read())
            })

        # fill IR-17 excel file
        cwf = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "IR-17-2015.xlsx")
        wb = load_workbook(cwf)
        ws1 = wb.get_sheet_by_name("IR17")  # Get sheet 1 in writeable copy
        for k, v in xls_dict["ir17"].iteritems():
            ws1[k] = v

        period = self.name.split("/")
        FILENAME = "IR-17-{}-{}.xlsx".format(period[0], period[1])
        wb.save("/tmp/{}".format(FILENAME))
        with open("/tmp/{}".format(FILENAME), "rb") as xls_file:
            self.write({
                'ir17_filename': FILENAME,
                'ir17_binary': base64.b64encode(xls_file.read())
            })
        out_inovice_url = "/web#id={}&view_type=form&model=account.invoice&action=203&menu_id=107"
        in_inovice_url = "/web#id={}&view_type=form&model=account.invoice&action=204&menu_id=107"
        if error_list:
            message = "<ul>"
            for ncf, errors in error_list.iteritems():
                message += "<li>{}</li><ul>".format(
                    errors[0][1] or "Factura invalida")
                for error in errors:
                    if error[0] in ("out_invoice", "out_refund"):
                        message += "<li><a target='_blank' href='{}'>{}</a></li>".format(
                            out_inovice_url.format(ncf), error[2])
                    else:
                        message += "<li><a target='_blank' href='{}'>{}</a></li>".format(
                            in_inovice_url.format(ncf), error[2])
                message += "</ul>"
            message += "</ul>"

            self.message_post(body=message)
            self.state = "error"
        else:
            self.message_post(body="Generado correctamente")
            self.state = "done"

    def generate_txt(self):
        company_fiscal_identificacion = re.sub("[^0-9]", "",
                                               self.company_id.vat)
        period = self.name.split("/")
        month = period[0]
        year = period[1]

        sale_path = '/tmp/607{}.txt'.format(company_fiscal_identificacion)
        sale_file = io.open(sale_path, 'w', encoding="utf-8", newline='\r\n')

        lines = []

        CANTIDAD_REGISTRO = "{:.2f}".format(len(self.sale_report)).zfill(12)
        TOTAL_MONTO_FACTURADO_FACTURAS = sum(
            [rec.MONTO_FACTURADO for rec in self.sale_report
             if rec.NUMERO_COMPROBANTE_MODIFICADO is False])
        TOTAL_MONTO_FACTURADO_NC = sum(
            [rec.MONTO_FACTURADO for rec in self.sale_report
             if rec.NUMERO_COMPROBANTE_MODIFICADO is not False])
        TOTAL_MONTO_FACTURADO = "{:.2f}".format(
            TOTAL_MONTO_FACTURADO_FACTURAS - TOTAL_MONTO_FACTURADO_NC).zfill(16)

        header = "607"
        header += company_fiscal_identificacion.rjust(11)
        header += str(year)
        header += str(month).zfill(2)
        header += CANTIDAD_REGISTRO
        header += TOTAL_MONTO_FACTURADO
        lines.append(header)

        for sale_line in self.sale_report:
            if not sale_line.NUMERO_COMPROBANTE_FISCAL:
                raise ValidationError(_(
                    "Por favor verifique los datos de la factura: {} en fecha: {}".format(
                        sale_line.NUMERO_COMPROBANTE_FISCAL, sale_line.FECHA_COMPROBANTE)))

            ln = ""
            ln += sale_line.RNC_CEDULA and sale_line.RNC_CEDULA.rjust(11) or "".rjust(11)
            ln += sale_line.TIPO_IDENTIFICACION
            ln += sale_line.NUMERO_COMPROBANTE_FISCAL.rjust(19)
            ln += sale_line.NUMERO_COMPROBANTE_MODIFICADO or "".rjust(19)
            ln += sale_line.FECHA_COMPROBANTE.replace("-", "")
            ln += "{:.2f}".format(sale_line.ITBIS_FACTURADO).zfill(12)
            ln += "{:.2f}".format(abs(sale_line.MONTO_FACTURADO)).zfill(12)
            lines.append(ln)

        for line in lines:
            sale_file.write(unicode(line) + "\n")

        sale_file.close()
        sale_file = open(sale_path, 'rb')
        sale_binary = base64.b64encode(sale_file.read())
        report_name = 'DGII_607_{}_{}{}.TXT'.format(
            company_fiscal_identificacion, str(year), str(month).zfill(2))
        self.write({'sale_binary': sale_binary, 'sale_filename': report_name})

        # 606
        pruchase_path = '/tmp/606{}.txt'.format(company_fiscal_identificacion)
        purchase_file = io.open(pruchase_path, 'w', encoding="utf-8",
                                newline='\r\n')
        lines = []

        CANTIDAD_REGISTRO = "{:.2f}".format(
            len(self.purchase_report)).zfill(12)
        TOTAL_MONTO_FACTURADO_FACTURAS = sum(
            [rec.MONTO_FACTURADO for rec in self.purchase_report
             if rec.NUMERO_COMPROBANTE_MODIFICADO is False])
        TOTAL_MONTO_FACTURADO_NC = sum(
            [rec.MONTO_FACTURADO for rec in self.purchase_report
             if rec.NUMERO_COMPROBANTE_MODIFICADO is not False])
        TOTAL_MONTO_FACTURADO = "{:.2f}".format(
            TOTAL_MONTO_FACTURADO_FACTURAS - TOTAL_MONTO_FACTURADO_NC).zfill(16)
        RETENCION_RENTA = "{:.2f}".format(sum([
            rec.RETENCION_RENTA for rec in self.purchase_report])).zfill(12)

        header = "606"
        header += company_fiscal_identificacion.rjust(11)
        header += str(year)
        header += str(month).zfill(2)
        header += CANTIDAD_REGISTRO
        header += TOTAL_MONTO_FACTURADO
        header += RETENCION_RENTA
        lines.append(header)

        for line in self.purchase_report:
            if not line.RNC_CEDULA or not line.NUMERO_COMPROBANTE_FISCAL:
                raise ValidationError(_(
                    "Por favor verifique la factura: {} de fecha {}".format(
                        line.NUMERO_COMPROBANTE_FISCAL, line.FECHA_COMPROBANTE)))
            ln = ""
            ln += line.RNC_CEDULA.rjust(11)
            ln += line.TIPO_IDENTIFICACION
            ln += line.TIPO_BIENES_SERVICIOS_COMPRADOS
            ln += line.NUMERO_COMPROBANTE_FISCAL and line.NUMERO_COMPROBANTE_FISCAL.rjust(
                19) or "".rjust(19)
            ln += line.NUMERO_COMPROBANTE_MODIFICADO or "".rjust(19)
            ln += line.FECHA_COMPROBANTE.replace("-", "")
            ln += line.FECHA_PAGO.replace("-",
                                          "") if line.FECHA_PAGO else "".rjust(8)
            ln += "{:.2f}".format(line.ITBIS_FACTURADO).zfill(12)
            ln += "{:.2f}".format(abs(line.ITBIS_RETENIDO)).zfill(12)
            ln += "{:.2f}".format(line.MONTO_FACTURADO).zfill(12)
            ln += "{:.2f}".format(line.RETENCION_RENTA).zfill(12)
            lines.append(ln)

        for line in lines:
            purchase_file.write(unicode(line) + "\n")

        purchase_file.close()
        purchase_file = open(pruchase_path, 'rb')
        purchase_binary = base64.b64encode(purchase_file.read())
        purchase_filename = 'DGII_606_{}_{}{}.TXT'.format(
            company_fiscal_identificacion, str(year), str(month).zfill(2))
        self.write({'purchase_binary': purchase_binary,
                    'purchase_filename': purchase_filename})

        # 608
        path = '/tmp/608{}.txt'.format(company_fiscal_identificacion)
        file = io.open(path, 'w', encoding="utf-8", newline='\r\n')
        lines = []

        header = "608"
        header += company_fiscal_identificacion.zfill(11)
        header += str(year)
        header += str(month).zfill(2)
        lines.append(header)

        for line in self.cancel_report:
            if not line.NUMERO_COMPROBANTE_FISCAL:
                raise ValidationError(_(
                    "Por favor verifique los datos de la factura: {} en fecha: {}".format(line.NUMERO_COMPROBANTE_FISCAL, line.FECHA_COMPROBANTE)))

            ln = ""
            ln += line.NUMERO_COMPROBANTE_FISCAL
            ln += line.FECHA_COMPROBANTE.replace("-", "")
            ln += "{}".format(line.TIPO_ANULACION).zfill(2)
            lines.append(ln)

        for line in lines:
            file.write(unicode(line) + "\n")

        file.close()
        file = open(path, 'rb')
        report = base64.b64encode(file.read())
        report_name = 'DGII_608_{}_{}{}.TXT'.format(
            company_fiscal_identificacion, str(year), str(month).zfill(2))
        self.write({'cancel_binary': report, 'cancel_filename': report_name})


class DgiiReportPurchaseLine(models.Model):
    _name = "dgii.report.purchase.line"

    dgii_report_id = fields.Many2one("dgii.report")
    LINE = fields.Integer("Linea")
    RNC_CEDULA = fields.Char("RNC", size=11)
    TIPO_IDENTIFICACION = fields.Char("Tipo ID", size=1)
    NUMERO_COMPROBANTE_FISCAL = fields.Char("NCF", size=19)
    NUMERO_COMPROBANTE_MODIFICADO = fields.Char("Afecta", size=19)
    FECHA_COMPROBANTE = fields.Date("Fecha")
    FECHA_PAGO = fields.Date("Pagado")

    TIPO_BIENES_SERVICIOS_COMPRADOS = fields.Char("Tipo", size=2)

    ITBIS_FACTURADO = fields.Float("ITBIS Facturado")
    ITBIS_RETENIDO = fields.Float("ITBIS Retenido")
    MONTO_FACTURADO = fields.Float("Monto Facturado")
    RETENCION_RENTA = fields.Float(u"Retención Renta")

    invoice = fields.Many2one("account.invoice", "NCF")
    number = fields.Char(related="invoice.number", string=" NCF")
    inv_partner = fields.Many2one(
        "res.partner", related="invoice.partner_id", string="Relacionado")
    affected_nvoice_id = fields.Many2one("account.invoice", "Afecta")
    nc = fields.Boolean()


class DgiiReportSaleLine(models.Model):
    _name = "dgii.report.sale.line"

    dgii_report_id = fields.Many2one("dgii.report")
    LINE = fields.Integer("Linea")
    RNC_CEDULA = fields.Char("RNC", size=11)
    TIPO_IDENTIFICACION = fields.Char("Tipo ID", size=1)
    NUMERO_COMPROBANTE_FISCAL = fields.Char("NCF", size=19)
    NUMERO_COMPROBANTE_MODIFICADO = fields.Char("Afecta", size=19)
    FECHA_COMPROBANTE = fields.Date("Fecha")
    FECHA_PAGO = fields.Date("Pagado")
    ITBIS_FACTURADO = fields.Float("ITBIS Facturado")
    MONTO_FACTURADO = fields.Float("Monto Facturado")

    invoice = fields.Many2one("account.invoice", "NCF")
    number = fields.Char(related="invoice.number", string="NCF")
    inv_partner = fields.Many2one("res.partner",
                                  related="invoice.partner_id")
    affected_nvoice_id = fields.Many2one("account.invoice", "Afecta")
    nc = fields.Boolean()


class DgiiCancelReportline(models.Model):
    _name = "dgii.cancel.report.line"

    dgii_report_id = fields.Many2one("dgii.report")
    LINE = fields.Integer("Linea")
    NUMERO_COMPROBANTE_FISCAL = fields.Char("NCF", size=19)
    FECHA_COMPROBANTE = fields.Date("Fecha")
    TIPO_ANULACION = fields.Char(u"Tipo de anulación", size=2)
    invoice = fields.Many2one("account.invoice", "Factura")


class DgiiExteriorReportline(models.Model):
    _name = "dgii.exterior.report.line"

    dgii_report_id = fields.Many2one("dgii.report")
    LINE = fields.Integer("Linea")
    TIPO_BIENES_SERVICIOS_COMPRADOS = fields.Char("Tipo", size=2)
    FECHA_COMPROBANTE = fields.Date("Fecha")
    FECHA_PAGO = fields.Date("Pagado")
    RETENCION_RENTA = fields.Float(u"Retención Renta")
    MONTO_FACTURADO = fields.Float("Monto Facturado")
    invoice = fields.Many2one("account.invoice", "Factura")
